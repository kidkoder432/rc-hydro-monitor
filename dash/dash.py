from time import time
from apscheduler.schedulers.background import BackgroundScheduler
from winsound import Beep
from datetime import datetime
import serial
import openpyxl as xl
from random import random

import eel
from eel import sleep

eel.init(r"C:\Users\HsuVictor2101\Downloads\car_monitor\esp_monitor\web")

FILENAME = r"C:\Users\HsuVictor2101\Downloads\car_monitor\esp_monitor\CarLog.xlsx"
PORT = "COM11"


# Colors
class colors:
    HEADER = "\033[95m"
    OKBLUE = "\033[94m"
    OKCYAN = "\033[96m"
    OKGREEN = "\033[32;1m"
    WARNING = "\033[93;1m"
    FAIL = "\033[91;1m"
    ENDC = "\033[0m"
    BOLD = "\033[1m"
    UNDERLINE = "\033[4m"


esp = serial.Serial(PORT, 115200, timeout=3)


team = input("What team are you on (1 or 2)?  ")

add = 0

if "1" in team:
    sheetName = "Car1"
    add = .3
elif "2" in team:
    sheetName = "Car2"
else:
    sheetName = "Test"

wb = xl.load_workbook(FILENAME)
sheet = wb[sheetName]

firstEmpty = sheet.max_row

timeouts = 0

idx = firstEmpty
lastRecvTime = time() - 2
startTime = time()
startV = 0
isVSet = False

startA = 0
isASet = False

v, a = 0, 0
volts, amps = [0], [0]


def r(a, b):
    if a > b:
        a, b = b, a
    return random() * (b - a) + a

def save(a, d):
    sch.shutdown()
    wb.save(FILENAME)
    print("Done")
    exit(0)

def t():
    try:
        print(colors.BOLD + datetime.now().strftime("%X"), colors.ENDC, end="\r")
        if time() - lastRecvTime > 10:
            red = True
        else:
            red = False
        eel.t(datetime.now().strftime("%X"), time() - lastRecvTime, red)

        eel.recv(volts[-1], amps[-1])
    except KeyboardInterrupt:
        save(0, 0)

def logData():
    global idx
    try:
        c = colors.OKGREEN
        if volts[-1] < 7:
            c = colors.FAIL
        elif amps[-1] < 0.3:
            c = colors.WARNING
        print(c, end="")
        print(
            datetime.now().strftime("%X"),
            volts[-1],
            "V,",
            amps[-1],
            "A",
            colors.ENDC,
        )
        for a, v in zip(amps, volts):

            sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), a, v])
            print(sheet.max_row)
    except KeyboardInterrupt:
        save(0, 0)





sch = BackgroundScheduler()
sch.add_job(t, trigger="interval", seconds=.1)
sch.add_job(logData, trigger="interval", seconds=10)
sch.start()

eel.start("index.html", mode="edge", block=False, close_callback=save)
sleep(1)

while True:

    try:
        data = esp.readline()[:-2].decode()  # f"##  {r(6, 9)}  {r(0, 3)}\t"  #

        if data.startswith("##"):
            lastRecvTime = time()

            data = data[4:]
            volts = []
            amps = []
            for d in data.split("\t")[:-1]:
                v, a = map(float, d.split("  "))
                if float(v) != 0.00:
                    volts.append(v + add)
                    amps.append(a)

            # # ONLY FOR TESTING
            # volts[-1] = 7.8 - (time() - startTime) * 0.004
            # amps[-1] = 2.4 - (time() - startTime) * 0.02

            if not isVSet:
                startV = volts[-1]
                isVSet = True

            if not isASet:
                startA = amps[-1]
                isASet = True

            if volts[-1] > 7:
                eel.setHTML("remvh", "Battery will die in approximately: <span id=\"remv\">0</span> sec.")
                eel.setRem("remv", startTime, time(), 8.6, volts[-1], 7.0)
            else:
                eel.setHTML("remvh", "<span class='r'>Battery is dead!</span>")

            if amps[-1] > 0:
                eel.setHTML("remah", "Hydrostik will die in approximately: <span id=\"rema\"></span> sec.")
                eel.setRem("rema", startTime, time(), 3.0, amps[-1], 0.0)
            else:
                eel.setHTML("remah", "<span class='r'>Hydrostik is dead!</span>")

        else:
            pass
            # print(data)


        sleep(max(0.01, (2 - (time() - lastRecvTime))))

    except (serial.SerialTimeoutException, serial.SerialException) as e:

        if "ClearCommError" in str(e):
            try:
                esp = serial.Serial(PORT, 115200, timeout=3)
            except serial.SerialException:
                wb.save(FILENAME)
                sch.pause()
                input(
                    "The Arduino is not plugged in. Please plug in the Arduino and try again. Press enter to continue"
                )
                sch.resume()
            continue

    except KeyboardInterrupt:
        save(0, 0)
